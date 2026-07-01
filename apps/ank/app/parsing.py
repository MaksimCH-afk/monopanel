"""Parsing of uploaded frequency / project files (Excel or CSV)."""
from __future__ import annotations

import csv
import io
import re

from openpyxl import load_workbook

# Column-name synonyms used when importing full projects from one workbook.
KEYWORD_SYNS = ("keyword", "ключ", "запрос", "phrase", "key word", "keywords", "фраза", "анкор")
VOLUME_SYNS = ("volume", "vol", "частотн", "частот", "freq", "wordstat", "ws", "показ", "трафик")
# KD / Keyword Difficulty and similar are intentionally ignored.

_DOMAIN_RE = re.compile(r"^(https?://)?([a-z0-9-]+\.)+[a-z]{2,}(/.*)?$", re.I)
# A (possibly malformed) leading scheme: http://, https://, http:/, http/, https:/ …
_SCHEME_RE = re.compile(r"^\s*https?[:/]+", re.I)
# A bare host (+ optional path), no scheme.
_BARE_DOMAIN_RE = re.compile(r"^([a-z0-9-]+\.)+[a-z]{2,}(/.*)?$", re.I)


def _rows_from_xlsx(content: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return _normalize_rows(ws)


def _normalize_rows(ws) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in row])
    return rows


def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Sniff delimiter (comma / semicolon / tab) from the first non-empty line.
    sample = next((ln for ln in text.splitlines() if ln.strip()), "")
    delimiter = ","
    for cand in [";", "\t", ","]:
        if cand in sample:
            delimiter = cand
            break
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[c.strip() for c in row] for row in reader]


def read_table(filename: str, content: bytes) -> list[list[str]]:
    """Read an uploaded file into a list of string rows (first sheet for Excel)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _rows_from_xlsx(content)
    return _rows_from_csv(content)


def _looks_like_header(cells: list[str]) -> bool:
    joined = " ".join(cells).lower()
    return any(w in joined for w in ("keyword", "ключ", "frequency", "freq", "частот", "volume", "объ", "vol"))


def rows_to_pairs(rows: list[list[str]]) -> list[tuple[str, float]]:
    """Turn raw string rows into ``[(keyword, frequency), ...]``.

    The first column is the keyword, the second (if any) the frequency/volume.
    A header row (if detected) is skipped. Rows without a keyword are ignored;
    a missing or non-numeric frequency defaults to 0.
    """
    out: list[tuple[str, float]] = []
    for i, cells in enumerate(rows):
        if not cells or not any(cells):
            continue
        if i == 0 and _looks_like_header(cells):
            continue
        keyword = cells[0].strip()
        if not keyword:
            continue
        freq = 0.0
        if len(cells) > 1:
            raw = cells[1].replace(",", ".").replace(" ", "")
            try:
                freq = float(raw)
            except ValueError:
                freq = 0.0
        out.append((keyword, freq))
    return out


def parse_frequency(filename: str, content: bytes) -> list[tuple[str, float]]:
    """Parse a single ``keyword | frequency`` table (first sheet for Excel)."""
    return rows_to_pairs(read_table(filename, content))


def parse_workbook_sheets(content: bytes) -> dict[str, list[tuple[str, float]]]:
    """Parse every sheet of an Excel workbook (simple first-two-columns mode)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result: dict[str, list[tuple[str, float]]] = {}
    for sheet_name in wb.sheetnames:
        result[sheet_name] = rows_to_pairs(_normalize_rows(wb[sheet_name]))
    return result


# --------------------------------------------------------------------------- #
# Smart project import (one workbook -> several projects)
# --------------------------------------------------------------------------- #
def _looks_like_domain(value: str) -> bool:
    v = value.strip()
    if not v or " " in v:
        return False
    # Tolerate a malformed scheme (http:/, http/, …) before checking the host.
    v = _SCHEME_RE.sub("", v)
    return bool(_BARE_DOMAIN_RE.match(v))


def normalize_domain(value: str) -> str:
    """Normalise a domain/URL to ``https://host/...``.

    Forces https, drops a leading ``www.``, ensures a trailing slash on the root
    form, and repairs malformed schemes — so all of these collapse to the same
    canonical URL::

        site.com/  •  https:/site.com/  •  http:/site.com  •  http/site.com/
        •  http://site.com/  •  https://www.site.com  ->  https://site.com/
    """
    v = value.strip()
    v = _SCHEME_RE.sub("", v)              # strip any (malformed) scheme
    v = re.sub(r"^www\.", "", v, flags=re.I)
    v = "https://" + v
    # Ensure the root form ends with a single slash.
    if "/" not in v.split("://", 1)[1]:
        v += "/"
    return v


def _find_header(rows: list[list[str]]) -> tuple[int | None, int | None, int | None]:
    """Locate the header row and the keyword / volume column indexes."""
    for idx, cells in enumerate(rows[:6]):
        low = [c.lower() for c in cells]
        kw_col = next((i for i, c in enumerate(low) if any(s in c for s in KEYWORD_SYNS)), None)
        if kw_col is None:
            continue
        vol_col = next((i for i, c in enumerate(low) if any(s in c for s in VOLUME_SYNS)), None)
        return idx, kw_col, vol_col
    return None, None, None


def _find_domains(rows: list[list[str]]) -> list[str]:
    """All distinct domains/URLs found in the sheet, normalised, in order.

    A sheet may list several domains that share one keyword set (e.g. mirror
    domains) — each becomes its own project.
    """
    seen: set[str] = set()
    out: list[str] = []
    for cells in rows:
        for c in cells:
            if _looks_like_domain(c):
                norm = normalize_domain(c)
                if norm not in seen:
                    seen.add(norm)
                    out.append(norm)
    return out


def _find_domain(rows: list[list[str]]) -> str | None:
    """The first domain/URL in the sheet (or ``None``)."""
    domains = _find_domains(rows)
    return domains[0] if domains else None


def _to_float(raw: str) -> float:
    try:
        return float(raw.replace(",", ".").replace(" ", ""))
    except ValueError:
        return 0.0


def parse_project_sheet_multi(rows: list[list[str]]) -> tuple[list[str], list[tuple[str, float]]]:
    """Parse one sheet into ``(domains, [(keyword, frequency), ...])``.

    Like :func:`parse_project_sheet` but returns *every* domain in the sheet —
    a sheet may list several domains that share one keyword set.
    """
    rows = [r for r in rows if any(r)]
    if not rows:
        return [], []
    header_idx, kw_col, vol_col = _find_header(rows)
    domains = _find_domains(rows)
    if header_idx is None or kw_col is None:
        return domains, []  # no keyword column -> not a project sheet

    pairs: list[tuple[str, float]] = []
    for cells in rows[header_idx + 1:]:
        if kw_col >= len(cells):
            continue
        keyword = cells[kw_col].strip()
        if not keyword or _looks_like_domain(keyword):
            continue
        freq = _to_float(cells[vol_col]) if (vol_col is not None and vol_col < len(cells)) else 0.0
        pairs.append((keyword, freq))
    return domains, pairs


def parse_project_sheet(rows: list[list[str]]) -> tuple[str | None, list[tuple[str, float]]]:
    """Parse one sheet into ``(domain, pairs)`` — the first domain only.

    Handles synonymous column names, columns in any position and a domain that
    can live in a header cell, a dedicated column or repeated in every row.
    Returns ``(None, [])`` for empty / summary / dashboard sheets.
    """
    domains, pairs = parse_project_sheet_multi(rows)
    return (domains[0] if domains else None), pairs


def parse_project_sheets(content: bytes) -> list[dict]:
    """Import projects from a workbook: one sheet = one or more projects.

    Returns a list of ``{"name": sheet_name, "domains": [str], "pairs": [...]}``.
    Each domain in a sheet becomes its own project (sharing the keyword set).
    Empty and summary sheets yield empty ``pairs`` and are skipped by the caller.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        domains, pairs = parse_project_sheet_multi(_normalize_rows(wb[sheet_name]))
        out.append({"name": sheet_name, "domains": domains, "pairs": pairs})
    return out


def parse_project_table(filename: str, content: bytes) -> dict:
    """Same as :func:`parse_project_sheets` but for a single CSV/Excel table."""
    rows = read_table(filename, content)
    domains, pairs = parse_project_sheet_multi(rows)
    return {"name": filename, "domains": domains, "pairs": pairs}
