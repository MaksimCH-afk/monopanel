"""Tests for the Excel export (columns, expansion, grouping, language, url type)."""
import io

from openpyxl import load_workbook

from app.excel_export import BASE_COLUMNS, anchor_type, build_workbook
from app.generator import GeneratedRow


def _rows():
    return {
        "Прогоны": [
            GeneratedRow(link_qty=3, url="https://betalice.com/", anchor="https://betalice.com/",
                         article_language="German", keyword="https://betalice.com/"),
            GeneratedRow(link_qty=2, url="https://betalice.com/", anchor="betalice",
                         article_language="German", keyword="betalice", is_keyword=True),
        ],
        "Внутренние страницы": [
            GeneratedRow(link_qty=1, url="https://betalice.com/boni/", anchor="betalice auszahlung",
                         article_language="German", keyword="betalice auszahlung"),
        ],
    }


def _load(content):
    return load_workbook(io.BytesIO(content))


def test_expanded_columns_and_row_count():
    wb = _load(build_workbook(_rows(), sprint="122", seo_specialist="Miles Nashwood", language="German"))
    ws = wb["Прогоны"]
    header = [c.value for c in ws[1]]
    assert header == BASE_COLUMNS + ["Article Language"]
    # 3 + 2 links expanded -> 5 data rows
    assert ws.max_row - 1 == 5
    first = [c.value for c in ws[2]]
    assert first[0] == "122"                       # Sprint
    assert first[1] == "Miles Nashwood"            # SEO Specialist
    assert first[2] == "betalice.com"              # Project = bare domain
    assert first[3] == "https://betalice.com/"     # Project Url
    assert first[4] == "Main Page"                 # URL Type
    assert first[-1] == "German"                   # Article Language


def test_grouped_has_link_qty_column():
    wb = _load(build_workbook(_rows(), grouped=True, language="German"))
    ws = wb["Прогоны"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Link Q-ty"
    # grouped: one row per anchor -> 2 rows
    assert ws.max_row - 1 == 2
    assert [c.value for c in ws[2]][0] == 3        # quantity in first column


def test_language_omitted_when_empty():
    wb = _load(build_workbook(_rows(), language=""))
    header = [c.value for c in wb["Прогоны"][1]]
    assert "Article Language" not in header


def test_internal_sheet_url_type_inner_page():
    wb = _load(build_workbook(_rows(), language="German"))
    ws = wb["Внутренние страницы"]
    row = [c.value for c in ws[2]]
    assert row[4] == "Inner Page"
    assert row[3] == "https://betalice.com/boni/"


def test_keyword_column_holds_top_keyword_on_every_row():
    kw_idx = BASE_COLUMNS.index("Keyword")
    wb = _load(build_workbook(_rows(), language="German"))
    # Every row in every sheet carries the project's most-used keyword.
    for sheet in ("Прогоны", "Внутренние страницы"):
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            assert [c.value for c in ws[r]][kw_idx] == "betalice"


def test_anchor_type_classification():
    # Branded: bare URL/domain or the brand/site name
    assert anchor_type("https://betalice.com/", is_keyword=False, brand="", domain="betalice.com") == "BD"
    assert anchor_type("betalice.com", is_keyword=False, brand="", domain="betalice.com") == "BD"
    assert anchor_type("AustriaWin24.at", is_keyword=False, brand="AustriaWin24", domain="austriawin24.at") == "BD"
    assert anchor_type("betalice", is_keyword=True, brand="betalice", domain="x.com") == "BD"
    # Exact match: a bare keyword, no brand inside
    assert anchor_type("online casino österreich", is_keyword=True, brand="betalice", domain="x.com") == "EM"
    assert anchor_type("klarna casino", is_keyword=True, brand="austriawin24", domain="austriawin24.at") == "EM"
    # Partial match: brand embedded in a phrase, or a non-keyword phrase
    assert anchor_type("betalice casino", is_keyword=True, brand="betalice", domain="x.com") == "PM"
    assert anchor_type("Casino bezahlen per Magenta", is_keyword=False, brand="austriawin24", domain="austriawin24.at") == "PM"


def test_anchor_type_column_in_export():
    at_idx = BASE_COLUMNS.index("Anchor Type")
    wb = _load(build_workbook(_rows(), language="German", brand="betalice"))
    ws = wb["Прогоны"]
    types = {[c.value for c in ws[r]][at_idx] for r in range(2, ws.max_row + 1)}
    # anchorless URL rows -> BD, "betalice" keyword (== brand) -> BD
    assert types <= {"BD", "EM", "PM"}
    assert "BD" in types
    # every Anchor Type cell is a 2-letter code
    for r in range(2, ws.max_row + 1):
        assert [c.value for c in ws[r]][at_idx] in ("BD", "EM", "PM")


def test_language_column_suppressed_by_flag():
    # Even with a language set, include_language=False drops the column.
    wb = _load(build_workbook(_rows(), language="German", include_language=False))
    header = [c.value for c in wb["Прогоны"][1]]
    assert "Article Language" not in header


def test_crowd_keyword_fallback_fills_column():
    kw_idx = BASE_COLUMNS.index("Keyword")
    # Crowd campaign: only anchorless rows (no keyword anchors).
    sheets = {
        "Крауд+сабмиты": [
            GeneratedRow(link_qty=3, url="https://x.com/", anchor="x.com",
                         article_language="", keyword="x.com"),
            GeneratedRow(link_qty=2, url="https://x.com/", anchor="https://x.com/",
                         article_language="", keyword="https://x.com/"),
        ],
    }
    wb = _load(build_workbook(sheets, keyword="best keyword"))
    ws = wb["Крауд+сабмиты"]
    for r in range(2, ws.max_row + 1):
        assert [c.value for c in ws[r]][kw_idx] == "best keyword"


def test_keyword_blank_for_fully_anchorless():
    kw_idx = BASE_COLUMNS.index("Keyword")
    sheets = {
        "Крауд+сабмиты": [
            GeneratedRow(link_qty=2, url="https://x.com/", anchor="https://x.com/",
                         article_language="", keyword="https://x.com/"),
        ],
    }
    wb = _load(build_workbook(sheets, language=""))
    ws = wb["Крауд+сабмиты"]
    assert [c.value for c in ws[2]][kw_idx] in (None, "")
